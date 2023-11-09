import time

import openpyxl
import requests

import config


class Crawler:
    def __init__(self):
        self.token = config.token
        self.file_name = "adoption.xlsx"
        self.file = openpyxl.load_workbook(self.file_name)
        self.sheet = self.file["repositories"]

        self.keyword_position = {
            "Espresso": 2,
            "UI Automator": 3,
            "Appium": 4
        }

    def search_code(self, keyword):
        for row in range(2, self.sheet.max_row + 1):
            cell = self.sheet.cell(row=row, column=1)
            owner = cell.value.split("/")[0]
            repo = cell.value.split("/")[1]
            response = requests.get(f"https://api.github.com/search/code?q={keyword}+repo:{owner}/{repo}",
                                    headers={"Authorization": f"token {self.token}"})
            data = response.json()
            print(f"repo: {cell.value} progress: {row - 1} / {self.sheet.max_row - 1} ")
            print(data)
            self.sheet.cell(row=row, column=self.keyword_position[keyword]).value = 0
            if data["total_count"] > 1:
                for item in data["items"]:
                    if "androidTest" in item["path"]:
                        self.sheet.cell(row=row, column=self.keyword_position[keyword]).value = 1
            self.file.save(self.file_name)
            time.sleep(10)

    def search_test(self, keyword, repo):
        owner = repo.split("/")[0]
        repo = repo.split("/")[1]
        response = requests.get(f"https://api.github.com/search/code?q={keyword}+repo:{owner}/{repo}",
                                headers={"Authorization": f"token {self.token}"})
        data = response.json()
        print(data)


if __name__ == '__main__':
    crawler = Crawler()
    crawler.search_code("Espresso")
    crawler.search_code("UI Automator")
    crawler.search_code("Appium")
    crawler.search_code("Espresso")
    crawler.search_code("UI Automator")
    crawler.search_code("Appium")
    crawler.search_code("Espresso")
    crawler.search_code("UI Automator")
    crawler.search_code("Appium")
